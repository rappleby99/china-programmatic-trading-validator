import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional, Callable
from enum import Enum


class Severity(Enum):
    ERROR = "ERROR"
    WARNING = "WARNING"


@dataclass
class ValidationError:
    row_num: int
    field_name_cn: str  # Chinese field name
    field_name_en: str  # English field name
    field_col: int      # Column number (1-based for Excel, or 0 for file-level)
    field_value: str
    message: str
    severity: Severity = Severity.ERROR
    account_name: str = ""
    client_code: str = ""

    def __str__(self):
        context = ""
        if self.account_name or self.client_code:
            parts = []
            if self.account_name:
                parts.append(f"Account: {self.account_name}")
            if self.client_code:
                parts.append(f"BCAN: {self.client_code}")
            context = f" [{', '.join(parts)}]"

        # Format: [ERROR] Row 1 [Account: X, BCAN: Y], Column 3 '账户名称' (account_name): message (value: 'X')
        field_info = f"Column {self.field_col} '{self.field_name_cn}' ({self.field_name_en})" if self.field_col > 0 else f"'{self.field_name_cn}'"
        return f"[{self.severity.value}] Row {self.row_num}{context}, {field_info}: {self.message} (value: '{self.field_value}')"


@dataclass
class RowValidationResult:
    """Result of validating a single row"""
    row_num: int
    account_name: str
    client_code: str
    is_valid: bool
    error_count: int = 0
    warning_count: int = 0


@dataclass
class FieldSpec:
    """Specification for a single field"""
    index: int
    name_cn: str
    name_en: str
    max_length: int
    required: bool = False  # Always required
    conditional_required: Optional[Callable] = None  # Conditionally required
    valid_values: Optional[list] = None  # Enumerated values (Chinese)
    multi_value: bool = False  # Can contain multiple values separated by ;
    max_multi_count: Optional[int] = None  # Maximum number of values if multi
    validator: Optional[Callable] = None  # Custom validation function
    description: str = ""


class SSEValidator:
    """Validator for SSE Programmatic Trading Reports"""

    # Enumerated values (must be in Chinese)
    REPORT_TYPES = ["首次", "变更", "停止使用"]
    YES_NO = ["是", "否"]
    YES_NO_EXEMPT = ["是", "否", "已申请豁免"]
    FUND_SOURCES = ["自有资金", "募集资金", "杠杆资金", "其他"]
    LEVERAGE_SOURCES = ["融资融券", "场外衍生品", "其他"]
    TRADING_PRODUCTS = ["股票", "基金"]
    STRATEGY_TYPES = [
        "指数增强策略", "市场中性策略", "多空灵活策略", "量化多头策略",
        "管理期货策略CTA", "参与新股发行策略", "量化套利策略", "日内回转策略", "其他"
    ]
    EXECUTION_METHODS = ["TWAP", "VWAP", "POV", "其他"]
    ORDER_RATES = ["500笔及以上", "300笔至499笔", "100笔至299笔", "100笔以下"]
    DAILY_ORDER_COUNTS = [
        "25000笔及以上", "20000笔至24999笔", "15000笔至19999笔",
        "10000笔至14999笔", "10000笔以下"
    ]

    # High frequency thresholds
    HIGH_FREQ_RATES = ["500笔及以上", "300笔至499笔"]
    HIGH_FREQ_DAILY = ["25000笔及以上", "20000笔至24999笔"]

    # Special value for consolidated reporting
    REPORTED_ELSEWHERE = "已在其他联交所参与者报告"
    EXEMPT_VALUE = "已申请豁免"

    def __init__(self):
        self.errors: list[ValidationError] = []
        self.row_results: list[RowValidationResult] = []
        self.exchange_type: Optional[str] = None  # SHANGHAI, SHENZHEN, or None
        self.submission_date: Optional[datetime] = None  # Date from filename (submission date)
        self.firm_id: Optional[str] = None  # Broker code from filename (5-digit FIRM_ID)
        self.field_specs = {}  # Will be built after exchange type is detected
        self.header_row = None
        self.current_row_context = {"account_name": "", "client_code": ""}

    @staticmethod
    def detect_exchange(filename: str) -> tuple[Optional[str], Optional[datetime], Optional[str], Optional[str]]:
        """
        Detect exchange type, submission date, and firm ID from filename.

        Returns:
            tuple: (exchange_type, submission_date, firm_id, error_message)
                - exchange_type: "SHANGHAI", "SHENZHEN", or None if invalid
                - submission_date: datetime object of the date in filename, or None if invalid
                - firm_id: 5-digit broker code from filename, or None if invalid
                - error_message: Error description if filename is invalid, None otherwise

        Expected formats:
            Shanghai: SH_PGTDRPT_<FIRM_ID>_<YYYYMMDD>.xlsx
            Shenzhen: SZ_PGTDRPT_<FIRM_ID>_<YYYYMMDD>.xlsx

        Where:
            FIRM_ID: 5-digit number with leading zeros (e.g., "09999")
            YYYYMMDD: Date in format YYYYMMDD (e.g., "20250805")
        """
        # Pattern: (SH|SZ)_PGTDRPT_<5 digits>_<8 digits>.xlsx
        pattern = r'^(SH|SZ)_PGTDRPT_(\d{5})_(\d{8})\.xlsx$'
        match = re.match(pattern, filename, re.IGNORECASE)

        if not match:
            return None, None, None, (
                f"Invalid filename format: '{filename}'. "
                f"Expected format: SH_PGTDRPT_<FIRM_ID>_<YYYYMMDD>.xlsx or "
                f"SZ_PGTDRPT_<FIRM_ID>_<YYYYMMDD>.xlsx "
                f"(e.g., SH_PGTDRPT_09999_20250805.xlsx)"
            )

        exchange_code = match.group(1).upper()
        firm_id = match.group(2)
        date_str = match.group(3)

        # Validate date format (basic check - YYYYMMDD should be valid)
        try:
            submission_date = datetime.strptime(date_str, '%Y%m%d')
        except ValueError:
            return None, None, None, f"Invalid date in filename: {date_str}. Expected format: YYYYMMDD"

        # Check that submission date is not in the future
        current_date = datetime.now()
        if submission_date > current_date:
            return None, None, None, f"Submission date in filename ({date_str}) cannot be in the future"

        # Map exchange code to full name
        if exchange_code == "SH":
            return "SHANGHAI", submission_date, firm_id, None
        elif exchange_code == "SZ":
            return "SHENZHEN", submission_date, firm_id, None
        else:
            return None, None, None, f"Unknown exchange code: {exchange_code}"

    def _build_field_specs(self, exchange_type: str = "SHANGHAI") -> dict[int, FieldSpec]:
        """Build field specifications based on exchange requirements

        Args:
            exchange_type: "SHANGHAI" or "SHENZHEN"
        """
        specs = {}

        # Determine if this is Shenzhen (has 序号 field at position 0)
        is_shenzhen = (exchange_type == "SHENZHEN")

        # Field index offset for Shenzhen (+1 due to 序号 column)
        offset = 1 if is_shenzhen else 0

        # Helper functions for conditional requirements (adjusted for offset)
        def req_if_first_or_change(row):
            return row.get(7 + offset, "") in ["首次", "变更"]

        def req_if_fund_source_other(row):
            # Shenzhen doesn't have "其他资金来源描述" field - never required
            if is_shenzhen:
                return False
            return "其他" in row.get(11 + offset, "")

        def req_if_has_leverage(row):
            return "杠杆资金" in row.get(11 + offset, "")

        def req_if_leverage_source_other(row):
            # Shenzhen doesn't have "其他杠杆资金来源描述" field - never required
            if is_shenzhen:
                return False
            return "其他" in row.get(15 + offset, "")

        def req_if_quantitative(row):
            return row.get(19 + offset, "") == "是"

        def req_if_main_strategy_other(row):
            # Shenzhen doesn't have "其他主策略类型" field - never required
            if is_shenzhen:
                return False
            return row.get(20 + offset, "") == "其他"

        def req_if_main_strategy_filled(row):
            return bool(row.get(20 + offset, "").strip())

        def req_if_sub_strategy_other(row):
            # Shenzhen doesn't have "其他辅策略类型" field - never required
            if is_shenzhen:
                return False
            return "其他" in row.get(23 + offset, "")

        def req_if_sub_strategy_filled(row):
            # For Shenzhen,辅策略类型 is at index 22, for Shanghai it's at 23
            idx = 22 if is_shenzhen else 23
            return bool(row.get(idx + offset, "").strip())

        def req_if_execution_other(row):
            # Shenzhen doesn't have "其他方式描述" field - never required
            if is_shenzhen:
                return False
            # For Shanghai, execution_method is at index 28
            return "其他" in row.get(28 + offset, "")

        def req_if_high_freq_no_exempt(row):
            # For Shenzhen: max_order_rate at 28, max_daily_orders at 29, upload_test_report at 37
            # For Shanghai: max_order_rate at 31, max_daily_orders at 32, upload_test_report at 40
            if is_shenzhen:
                rate_idx, daily_idx, upload_idx = 28, 29, 37
            else:
                rate_idx, daily_idx, upload_idx = 31, 32, 40

            rate = row.get(rate_idx + offset, "")
            daily = row.get(daily_idx + offset, "")
            is_high_freq = rate in self.HIGH_FREQ_RATES or daily in self.HIGH_FREQ_DAILY
            is_exempt = row.get(upload_idx + offset, "") == "已申请豁免"
            return is_high_freq and not is_exempt and req_if_first_or_change(row)

        def req_if_qfii_exemption(row):
            """QFII code required when using order-splitting exemption for high-freq reporting"""
            # For Shenzhen: max_order_rate at 28, max_daily_orders at 29, upload_test_report at 37
            # For Shanghai: max_order_rate at 31, max_daily_orders at 32, upload_test_report at 40
            if is_shenzhen:
                rate_idx, daily_idx, upload_idx = 28, 29, 37
            else:
                rate_idx, daily_idx, upload_idx = 31, 32, 40

            rate = row.get(rate_idx + offset, "")
            daily = row.get(daily_idx + offset, "")
            is_high_freq = rate in self.HIGH_FREQ_RATES or daily in self.HIGH_FREQ_DAILY
            upload_report = row.get(upload_idx + offset, "")
            # Required if high-freq but not uploading report (implying order-splitting exemption)
            # Based on Excel comment: QFII investors using order-splitting to avoid high-freq extra reporting
            return is_high_freq and upload_report == "否" and req_if_first_or_change(row)

        # Field definitions (0-indexed to match CSV columns after header)
        # Structure: (index, name_cn, name_en, max_length, required, conditional_required, valid_values, multi_value, max_multi_count)

        if is_shenzhen:
            # Shenzhen field definitions (38 fields with 序号 at position 0)
            field_defs = [
                # Shenzhen-specific: Sequence Number
                (0, "序号", "sequence_num", 10, False, None, None, False, None),

                # Basic Information (基本信息)
                (1, "联交所参与者名称", "ep_name", 100, True, None, None, False, None),
                (2, "经纪商代码", "broker_code", 5, True, None, None, False, None),
                (3, "账户名称", "account_name", 200, True, None, None, False, None),
                (4, "证件号码", "id_number", 80, False, req_if_first_or_change, None, False, None),
                (5, "产品编码（选填）", "product_code", 50, False, None, None, False, None),
                (6, "深市券商客户编码", "client_code", 10, True, None, None, False, None),
                (7, "产品管理机构名称", "fund_manager", 200, False, None, None, False, None),
                (8, "报告类型（首次/变更/停止使用）", "report_type", 6, True, None, self.REPORT_TYPES, False, None),
                (9, "报告日期", "report_date", 8, True, None, None, False, None),

                # Fund Information (资金信息)
                (10, "是否选取一家联交所参与者集中填报资金信息", "consolidated_reporting", 1, False, req_if_first_or_change, self.YES_NO, False, None),
                (11, "账户资金规模（人民币，万元）", "fund_size", 30, False, req_if_first_or_change, None, False, None),
                (12, "账户资金来源", "fund_sources", 30, False, req_if_first_or_change, self.FUND_SOURCES, True, None),
                # Note: Shenzhen doesn't have "其他资金来源描述" field (index 12 in Shanghai)
                (13, "资金来源占比（%）", "fund_source_ratio", 50, False, req_if_first_or_change, None, False, None),
                (14, "杠杆资金规模（人民币，万元）", "leverage_size", 30, False, req_if_first_or_change, None, False, None),
                (15, "杠杆资金来源", "leverage_sources", 30, False, req_if_has_leverage, self.LEVERAGE_SOURCES, True, None),
                # Note: Shenzhen doesn't have "其他杠杆资金来源描述" field (index 16 in Shanghai)
                (16, "杠杆率（%）", "leverage_ratio", 20, False, req_if_first_or_change, None, False, None),

                # Trading Information (交易信息)
                (17, "交易品种", "trading_products", 20, False, req_if_first_or_change, self.TRADING_PRODUCTS, True, None),
                (18, "是否量化交易", "is_quantitative", 1, False, req_if_first_or_change, self.YES_NO, False, None),
                (19, "主策略类型", "main_strategy", 20, False, req_if_quantitative, self.STRATEGY_TYPES, False, None),
                # Note: Shenzhen doesn't have "其他主策略类型" field (index 21 in Shanghai)
                (20, "主策略概述", "main_strategy_desc", 500, False, req_if_main_strategy_filled, None, False, None),
                (21, "辅策略类型", "sub_strategy", 50, False, None, self.STRATEGY_TYPES, True, 2),
                # Note: Shenzhen doesn't have "其他辅策略类型" field (index 24 in Shanghai)
                (22, "辅策略概述", "sub_strategy_desc", 500, False, req_if_sub_strategy_filled, None, False, None),
                (23, "期货市场账户名称（选填）", "futures_account_name", 200, False, None, None, True, None),
                (24, "期货市场账户代码（选填）", "futures_account_code", 300, False, None, None, True, None),
                (25, "交易指令执行方式", "execution_method", 50, False, req_if_first_or_change, self.EXECUTION_METHODS, True, None),
                # Note: Shenzhen doesn't have "其他方式描述" field (index 29 in Shanghai)
                (26, "交易指令执行方式概述", "execution_desc", 500, False, req_if_first_or_change, None, False, None),
                (27, "账户最高申报速率（笔/秒）", "max_order_rate", 20, False, req_if_first_or_change, self.ORDER_RATES, False, None),
                (28, "账户单日最高申报笔数（笔）", "max_daily_orders", 20, False, req_if_first_or_change, self.DAILY_ORDER_COUNTS, False, None),

                # Software Information (交易软件信息)
                (29, "程序化交易软件名称及版本号", "software_name", 200, False, req_if_first_or_change, None, True, None),
                (30, "程序化交易软件开发主体", "software_developer", 200, False, req_if_first_or_change, None, True, None),

                # Other Information (其他)
                (31, "高频交易系统服务器所在地", "hft_server_location", 100, False, req_if_high_freq_no_exempt, None, False, None),
                (32, "联交所参与者联络人（选填）", "ep_contact", 80, False, None, None, False, None),
                (33, "联系方式（选填）", "ep_contact_info", 80, False, None, None, False, None),
                (34, "投资者相关业务负责人（选填）", "investor_contact", 80, False, None, None, False, None),
                (35, "联系方式（选填）", "investor_contact_info", 80, False, None, None, False, None),
                (36, "是否提交测试报告及应急方案", "upload_test_report", 5, False, req_if_first_or_change, self.YES_NO_EXEMPT, False, None),
                (37, "合格境外投资者编码", "qfii_code", 50, False, req_if_qfii_exemption, None, False, None),
            ]
        else:
            # Shanghai field definitions (42 fields, original)
            field_defs = [
                # Basic Information (基本信息)
                (0, "联交所参与者名称", "ep_name", 100, True, None, None, False, None),
                (1, "经纪商代码", "broker_code", 5, True, None, None, False, None),
                (2, "账户名称", "account_name", 200, True, None, None, False, None),
                (3, "证件号码", "id_number", 80, False, req_if_first_or_change, None, False, None),
                (4, "产品编码（选填）", "product_code", 50, False, None, None, False, None),
                (5, "券商客户编码", "client_code", 10, True, None, None, False, None),
                (6, "产品管理机构名称", "fund_manager", 200, False, None, None, False, None),
                (7, "报告类型", "report_type", 6, True, None, self.REPORT_TYPES, False, None),
                (8, "报告日期", "report_date", 8, True, None, None, False, None),

                # Fund Information (资金信息)
                (9, "是否选取一家联交所参与者集中填报资金信息", "consolidated_reporting", 1, False, req_if_first_or_change, self.YES_NO, False, None),
                (10, "账户资金规模（人民币，万元）", "fund_size", 30, False, req_if_first_or_change, None, False, None),
                (11, "账户资金来源", "fund_sources", 30, False, req_if_first_or_change, self.FUND_SOURCES, True, None),
                (12, "其他资金来源描述", "other_fund_desc", 200, False, req_if_fund_source_other, None, False, None),
                (13, "资金来源占比", "fund_source_ratio", 50, False, req_if_first_or_change, None, False, None),
                (14, "杠杆资金规模（人民币，万元）", "leverage_size", 30, False, req_if_first_or_change, None, False, None),
                (15, "杠杆资金来源", "leverage_sources", 30, False, req_if_has_leverage, self.LEVERAGE_SOURCES, True, None),
                (16, "其他杠杆资金来源描述", "other_leverage_desc", 200, False, req_if_leverage_source_other, None, False, None),
                (17, "杠杆率（%）", "leverage_ratio", 20, False, req_if_first_or_change, None, False, None),

                # Trading Information (交易信息)
                (18, "交易品种", "trading_products", 20, False, req_if_first_or_change, self.TRADING_PRODUCTS, True, None),
                (19, "是否量化交易", "is_quantitative", 1, False, req_if_first_or_change, self.YES_NO, False, None),
                (20, "主策略类型", "main_strategy", 20, False, req_if_quantitative, self.STRATEGY_TYPES, False, None),
                (21, "其他主策略类型", "other_main_strategy", 200, False, req_if_main_strategy_other, None, False, None),
                (22, "主策略概述", "main_strategy_desc", 500, False, req_if_main_strategy_filled, None, False, None),
                (23, "辅策略类型", "sub_strategy", 50, False, None, self.STRATEGY_TYPES, True, 2),
                (24, "其他辅策略类型", "other_sub_strategy", 200, False, req_if_sub_strategy_other, None, False, None),
                (25, "辅策略概述", "sub_strategy_desc", 500, False, req_if_sub_strategy_filled, None, False, None),
                (26, "期货市场账户名称（选填）", "futures_account_name", 200, False, None, None, True, None),
                (27, "期货市场账户代码（选填）", "futures_account_code", 300, False, None, None, True, None),
                (28, "交易指令执行方式", "execution_method", 50, False, req_if_first_or_change, self.EXECUTION_METHODS, True, None),
                (29, "其他方式描述", "other_execution_desc", 500, False, req_if_execution_other, None, False, None),
                (30, "指令执行方式概述", "execution_desc", 500, False, req_if_first_or_change, None, False, None),
                (31, "账户最高申报速率", "max_order_rate", 20, False, req_if_first_or_change, self.ORDER_RATES, False, None),
                (32, "账户单日最高申报笔数", "max_daily_orders", 20, False, req_if_first_or_change, self.DAILY_ORDER_COUNTS, False, None),

                # Software Information (交易软件信息)
                (33, "程序化交易软件名称及版本号", "software_name", 200, False, req_if_first_or_change, None, True, None),
                (34, "程序化交易软件开发主体", "software_developer", 200, False, req_if_first_or_change, None, True, None),

                # Other Information (其他)
                (35, "高频交易系统服务器所在地", "hft_server_location", 100, False, req_if_high_freq_no_exempt, None, False, None),
                (36, "联交所参与者联络人（选填）", "ep_contact", 80, False, None, None, False, None),
                (37, "联交所参与者联络人联系方式（选填）", "ep_contact_info", 80, False, None, None, False, None),
                (38, "投资者相关业务负责人（选填）", "investor_contact", 80, False, None, None, False, None),
                (39, "投资者相关业务负责人联系方式（选填）", "investor_contact_info", 80, False, None, None, False, None),
                (40, "是否上传测试报告及应急方案", "upload_test_report", 5, False, req_if_first_or_change, self.YES_NO_EXEMPT, False, None),
                (41, "合格境外投资者编码", "qfii_code", 50, False, req_if_qfii_exemption, None, False, None),
            ]

        for idx, name_cn, name_en, max_len, required, cond_req, valid_vals, multi, max_multi in field_defs:
            specs[idx] = FieldSpec(
                index=idx,
                name_cn=name_cn,
                name_en=name_en,
                max_length=max_len,
                required=required,
                conditional_required=cond_req,
                valid_values=valid_vals,
                multi_value=multi,
                max_multi_count=max_multi
            )

        return specs

    def add_error(self, row_num: int, field_name_cn: str, field_name_en: str, field_col: int,
                  value: str, message: str, severity: Severity = Severity.ERROR):
        """Add a validation error with complete field information"""
        self.errors.append(ValidationError(
            row_num=row_num,
            field_name_cn=field_name_cn,
            field_name_en=field_name_en,
            field_col=field_col,
            field_value=value,
            message=message,
            severity=severity,
            account_name=self.current_row_context.get("account_name", ""),
            client_code=self.current_row_context.get("client_code", "")
        ))

    def add_error_for_field(self, row_num: int, field_idx: int, value: str, message: str, severity: Severity = Severity.ERROR):
        """Add error using field index to look up field spec"""
        spec = self.field_specs.get(field_idx)
        if spec:
            self.add_error(row_num, spec.name_cn, spec.name_en, field_idx + 1, value, message, severity)
        else:
            # Fallback for fields without spec
            self.add_error(row_num, f"Field {field_idx + 1}", f"field_{field_idx}", field_idx + 1, value, message, severity)

    def get_field_idx_by_chinese_name(self, name_cn: str) -> int:
        """Get field index by Chinese field name"""
        for idx, spec in self.field_specs.items():
            if spec.name_cn == name_cn:
                return idx
        return -1  # Not found

    def _get_field_idx(self, field_en_name: str) -> int:
        """Get field index by English field name"""
        for idx, spec in self.field_specs.items():
            if spec.name_en == field_en_name:
                return idx
        return -1

    def validate_broker_code(self, row_num: int, value: str) -> bool:
        """Validate broker code is exactly 5 digits and matches filename FIRM_ID"""
        if not value:
            return True  # Handled by required check

        broker_code_idx = self._get_field_idx("broker_code")
        if broker_code_idx == -1:
            return True

        # Check format (5 digits)
        if not re.match(r'^\d{5}$', value):
            self.add_error_for_field(row_num, broker_code_idx, value, "Must be exactly 5 digits")
            return False

        # Check that broker code matches the FIRM_ID from filename
        if self.firm_id and value != self.firm_id:
            self.add_error_for_field(row_num, broker_code_idx, value,
                          f"Broker code must match filename FIRM_ID ({self.firm_id})")
            return False

        return True

    def validate_client_code(self, row_num: int, value: str) -> bool:
        """Validate client code is 3-10 characters"""
        if not value:
            return True

        client_code_idx = self._get_field_idx("client_code")
        if client_code_idx == -1:
            return True

        if len(value) < 3 or len(value) > 10:
            self.add_error_for_field(row_num, client_code_idx, value, "Must be 3-10 characters")
            return False
        return True

    def validate_date(self, row_num: int, value: str) -> bool:
        """Validate date format YYYYMMDD and ensure not later than submission date"""
        if not value:
            return True

        report_date_idx = self._get_field_idx("report_date")
        if report_date_idx == -1:
            return True

        if not re.match(r'^\d{8}$', value):
            self.add_error_for_field(row_num, report_date_idx, value, "Must be in YYYYMMDD format")
            return False
        try:
            report_date = datetime.strptime(value, '%Y%m%d')
            # Check that report date is not later than submission date (from filename)
            # Per HKEX spec: "第 n 行[报告日期]晚于上传日期" - Report date cannot be later than upload date
            if self.submission_date and report_date > self.submission_date:
                self.add_error_for_field(row_num, report_date_idx, value,
                              f"Report date cannot be later than submission date ({self.submission_date.strftime('%Y%m%d')})")
                return False
        except ValueError:
            self.add_error_for_field(row_num, report_date_idx, value, "Invalid date")
            return False
        return True

    def validate_numeric(self, row_num: int, field_idx: int, value: str, allow_negative: bool = False) -> bool:
        """Validate numeric field (allows up to 2 decimal places)"""
        if not value or value == self.REPORTED_ELSEWHERE:
            return True
        try:
            num = float(value)
            if not allow_negative and num < 0:
                self.add_error_for_field(row_num, field_idx, value, "Must be non-negative")
                return False
            # Check decimal places
            if '.' in value:
                decimals = len(value.split('.')[1])
                if decimals > 2:
                    self.add_error_for_field(row_num, field_idx, value, "Maximum 2 decimal places allowed")
                    return False
        except ValueError:
            self.add_error_for_field(row_num, field_idx, value, "Must be a valid number")
            return False
        return True

    def validate_leverage_ratio(self, row_num: int, value: str, has_leverage: bool) -> bool:
        """Validate leverage ratio"""
        if not value or value == self.REPORTED_ELSEWHERE:
            return True

        leverage_ratio_idx = self._get_field_idx("leverage_ratio")
        if leverage_ratio_idx == -1:
            return True

        try:
            ratio = float(value)
            if ratio < 100:
                self.add_error_for_field(row_num, leverage_ratio_idx, value, "Must be >= 100 (leverage ratio cannot be less than 100%)")
                return False
            if not has_leverage and ratio != 100:
                self.add_error_for_field(row_num, leverage_ratio_idx, value, "Must be 100 when no leverage funds")
                return False
            if has_leverage and ratio <= 100:
                self.add_error_for_field(row_num, leverage_ratio_idx, value, "Must be > 100 when leverage funds exist")
                return False
        except ValueError:
            self.add_error_for_field(row_num, leverage_ratio_idx, value, "Must be a valid number")
            return False
        return True

    def validate_fund_source_ratio(self, row_num: int, ratio_value: str, sources_value: str) -> bool:
        """Validate fund source ratio matches sources and sums to 100%"""
        if not ratio_value or ratio_value == self.REPORTED_ELSEWHERE:
            return True
        if not sources_value or sources_value == self.REPORTED_ELSEWHERE:
            return True

        fund_source_ratio_idx = self._get_field_idx("fund_source_ratio")
        if fund_source_ratio_idx == -1:
            return True

        # Parse sources
        sources = [s.strip() for s in sources_value.split(';') if s.strip()]

        # Parse ratios (format: "自有资金80%;募集资金20%")
        ratio_pattern = r'([^;]+?)(\d+(?:\.\d+)?)\s*%'
        matches = re.findall(ratio_pattern, ratio_value)

        if not matches:
            self.add_error_for_field(row_num, fund_source_ratio_idx, ratio_value,
                          "Invalid format. Expected: '来源1XX%;来源2XX%'")
            return False

        ratio_sources = []
        total = 0
        for source, pct in matches:
            source = source.strip()
            ratio_sources.append(source)
            total += float(pct)

        # Check sources match
        for src in sources:
            if src not in ratio_sources:
                self.add_error_for_field(row_num, fund_source_ratio_idx, ratio_value,
                              f"Missing ratio for source: {src}")
                return False

        # Check total is 100%
        if abs(total - 100) > 0.01:
            self.add_error_for_field(row_num, fund_source_ratio_idx, ratio_value,
                          f"Ratios must sum to 100% (current: {total}%)")
            return False

        return True

    def validate_leverage_funds(self, row_num: int, leverage_size: str, fund_size: str, has_leverage: bool) -> bool:
        """Validate leverage fund size is consistent"""
        if leverage_size == self.REPORTED_ELSEWHERE or fund_size == self.REPORTED_ELSEWHERE:
            return True

        leverage_size_idx = self._get_field_idx("leverage_size")
        if leverage_size_idx == -1:
            return True

        try:
            lev_val = float(leverage_size) if leverage_size else 0
            fund_val = float(fund_size) if fund_size else 0

            if not has_leverage and lev_val != 0:
                self.add_error_for_field(row_num, leverage_size_idx, leverage_size,
                              "Must be 0 when leverage not in fund sources")
                return False

            if has_leverage and lev_val <= 0:
                self.add_error_for_field(row_num, leverage_size_idx, leverage_size,
                              "Must be > 0 when leverage is in fund sources")
                return False

            if lev_val > fund_val and fund_val > 0:
                self.add_error_for_field(row_num, leverage_size_idx, leverage_size,
                              f"Cannot exceed total fund size ({fund_size})")
                return False

        except ValueError:
            pass  # Handled by numeric validation

        return True

    def validate_multi_value_field(self, row_num: int, field_spec: FieldSpec, value: str) -> bool:
        """Validate multi-value fields (with or without enumerated values)"""
        if not value or value == self.REPORTED_ELSEWHERE:
            return True

        values = [v.strip() for v in value.split(';') if v.strip()]

        # Check for duplicates
        if len(values) != len(set(values)):
            self.add_error(row_num, field_spec.name_cn, field_spec.name_en, field_spec.index + 1,
                          value, "Duplicate values not allowed")
            return False

        # Check max count
        if field_spec.max_multi_count and len(values) > field_spec.max_multi_count:
            self.add_error(row_num, field_spec.name_cn, field_spec.name_en, field_spec.index + 1,
                          value, f"Maximum {field_spec.max_multi_count} values allowed")
            return False

        # Check each value is valid (only if field has enumerated values)
        if field_spec.valid_values:
            for v in values:
                if v not in field_spec.valid_values:
                    self.add_error(row_num, field_spec.name_cn, field_spec.name_en, field_spec.index + 1,
                                  value, f"Invalid value '{v}'. Must be one of: {', '.join(field_spec.valid_values)}")
                    return False

        # Check for whitespace in values
        for v in values:
            if v != v.strip() or '  ' in v:
                self.add_error(row_num, field_spec.name_cn, field_spec.name_en, field_spec.index + 1,
                              value, "Values must not contain leading/trailing spaces or line breaks")
                return False

        return True

    def validate_high_freq_requirements(self, row_num: int, row: dict) -> bool:
        """Validate high-frequency trading requirements"""
        rate_idx = self._get_field_idx("max_order_rate")
        daily_idx = self._get_field_idx("max_daily_orders")
        server_idx = self._get_field_idx("hft_server_location")
        upload_idx = self._get_field_idx("upload_test_report")

        if rate_idx == -1 or daily_idx == -1 or server_idx == -1 or upload_idx == -1:
            return True

        rate = row.get(rate_idx, "")
        daily = row.get(daily_idx, "")
        server_location = row.get(server_idx, "")
        upload_report = row.get(upload_idx, "")

        is_high_freq = rate in self.HIGH_FREQ_RATES or daily in self.HIGH_FREQ_DAILY

        if is_high_freq:
            if upload_report == "否":
                self.add_error_for_field(row_num, upload_idx, upload_report,
                              "High-frequency accounts must upload test report or apply for exemption")
                return False

            if upload_report == "是" and not server_location:
                self.add_error_for_field(row_num, server_idx, server_location,
                              "Required for high-frequency trading accounts", Severity.WARNING)

            if upload_report == "已申请豁免" and server_location and server_location != "已申请豁免":
                self.add_error_for_field(row_num, server_idx, server_location,
                              "Should be '已申请豁免' when applying for exemption", Severity.WARNING)

        return True

    def validate_row(self, row_num: int, row_data: list) -> bool:
        """Validate a single row of data"""
        # Determine max columns based on exchange type (Shenzhen has 38, Shanghai has 42)
        max_cols = 38 if self.exchange_type == "SHENZHEN" else 42

        # Build row dict for easier access
        row = {i: (row_data[i].strip() if i < len(row_data) else "") for i in range(max_cols)}

        # Get field indices dynamically based on exchange type
        account_name_idx = self._get_field_idx("account_name")
        client_code_idx = self._get_field_idx("client_code")
        report_type_idx = self._get_field_idx("report_type")
        ep_name_idx = self._get_field_idx("ep_name")
        broker_code_idx = self._get_field_idx("broker_code")
        report_date_idx = self._get_field_idx("report_date")

        # Set current row context for error messages
        self.current_row_context = {
            "account_name": row.get(account_name_idx, "") if account_name_idx != -1 else "",
            "client_code": row.get(client_code_idx, "") if client_code_idx != -1 else ""
        }

        valid = True
        report_type = row.get(report_type_idx, "") if report_type_idx != -1 else ""
        is_stop = report_type == "停止使用"

        # For "停止使用", only basic fields required
        if is_stop:
            required_field_names = ["ep_name", "broker_code", "account_name", "client_code", "report_date"]
            for field_name in required_field_names:
                idx = self._get_field_idx(field_name)
                spec = self.field_specs.get(idx) if idx != -1 else None
                if spec and not row.get(idx, ""):
                    self.add_error(row_num, spec.name_cn, spec.name_en, spec.index + 1, "", "Required field")
                    valid = False
            return valid

        # Validate each field
        for idx, spec in self.field_specs.items():
            value = row.get(idx, "")

            # Check length
            if value and len(value) > spec.max_length:
                display_value = value[:50] + "..." if len(value) > 50 else value
                self.add_error(row_num, spec.name_cn, spec.name_en, spec.index + 1, display_value,
                              f"Exceeds maximum length of {spec.max_length}")
                valid = False

            # Check required
            is_required = spec.required
            if spec.conditional_required:
                is_required = is_required or spec.conditional_required(row)

            if is_required and not value:
                self.add_error(row_num, spec.name_cn, spec.name_en, spec.index + 1, "", "Required field")
                valid = False
                continue

            # Skip further validation if empty
            if not value:
                continue

            # Check multi-value fields
            if spec.multi_value:
                if not self.validate_multi_value_field(row_num, spec, value):
                    valid = False
            # Check enumerated values (for non-multi-value fields)
            elif spec.valid_values:
                if value not in spec.valid_values and value != self.REPORTED_ELSEWHERE:
                    self.add_error(row_num, spec.name_cn, spec.name_en, spec.index + 1, value,
                                  f"Must be one of: {', '.join(spec.valid_values)}")
                    valid = False

        # Field-specific validations using dynamic lookup
        broker_val = row.get(broker_code_idx, "") if broker_code_idx != -1 else ""
        if not self.validate_broker_code(row_num, broker_val):
            valid = False

        client_val = row.get(client_code_idx, "") if client_code_idx != -1 else ""
        if not self.validate_client_code(row_num, client_val):
            valid = False

        date_val = row.get(report_date_idx, "") if report_date_idx != -1 else ""
        if not self.validate_date(row_num, date_val):
            valid = False

        # Numeric validations
        fund_size_idx = self._get_field_idx("fund_size")
        if fund_size_idx != -1:
            if not self.validate_numeric(row_num, fund_size_idx, row.get(fund_size_idx, "")):
                valid = False

        leverage_size_idx = self._get_field_idx("leverage_size")
        if leverage_size_idx != -1:
            if not self.validate_numeric(row_num, leverage_size_idx, row.get(leverage_size_idx, "")):
                valid = False

        # Leverage-related validations
        fund_sources_idx = self._get_field_idx("fund_sources")
        has_leverage = "杠杆资金" in row.get(fund_sources_idx, "") if fund_sources_idx != -1 else False

        leverage_ratio_idx = self._get_field_idx("leverage_ratio")
        if leverage_ratio_idx != -1:
            if not self.validate_leverage_ratio(row_num, row.get(leverage_ratio_idx, ""), has_leverage):
                valid = False

        if fund_size_idx != -1 and leverage_size_idx != -1:
            if not self.validate_leverage_funds(row_num, row.get(leverage_size_idx, ""), row.get(fund_size_idx, ""), has_leverage):
                valid = False

        # Fund source ratio validation
        fund_source_ratio_idx = self._get_field_idx("fund_source_ratio")
        if fund_source_ratio_idx != -1 and fund_sources_idx != -1:
            if not self.validate_fund_source_ratio(row_num, row.get(fund_source_ratio_idx, ""), row.get(fund_sources_idx, "")):
                valid = False

        # High-frequency trading validations
        if not self.validate_high_freq_requirements(row_num, row):
            valid = False

        return valid

    def validate_file(self, file_path: str, original_filename: Optional[str] = None) -> tuple[bool, list[ValidationError]]:
        """Validate an entire Excel file

        Args:
            file_path: Path to the Excel file to validate
            original_filename: Optional original filename (used when file_path is a temp file)
        """
        self.errors = []
        path = Path(file_path)

        if not path.exists():
            self.add_error(0, "File", "file", 0, str(file_path), "File not found")
            return False, self.errors

        # Only support Excel files
        if path.suffix.lower() != '.xlsx':
            self.add_error(0, "File", "file", 0, str(file_path), "Only Excel files (.xlsx) are supported")
            return False, self.errors

        # Detect exchange type, submission date, and firm ID from filename
        # Use original filename if provided (for uploaded files), otherwise use actual file path
        filename = original_filename if original_filename else path.name
        exchange_type, submission_date, firm_id, filename_error = self.detect_exchange(filename)

        if filename_error:
            self.add_error(0, "Filename", "filename", 0, filename, filename_error)
            return False, self.errors

        self.exchange_type = exchange_type
        self.submission_date = submission_date
        self.firm_id = firm_id

        # Build field specifications based on detected exchange type
        self.field_specs = self._build_field_specs(exchange_type)

        # Read Excel data
        rows = self._read_xlsx(file_path)

        if not rows:
            self.add_error(0, "File", "file", 0, str(file_path), "No data rows found")
            return False, self.errors

        # Track client codes for duplicate check
        client_codes = {}

        # Validate each row
        for row_num, row_data in enumerate(rows, start=1):
            # Track errors before validation
            errors_before = len([e for e in self.errors if e.severity == Severity.ERROR and e.row_num == row_num])
            warnings_before = len([e for e in self.errors if e.severity == Severity.WARNING and e.row_num == row_num])

            self.validate_row(row_num, row_data)

            # Check for duplicate client codes using dynamic index
            client_code_idx = self._get_field_idx("client_code")
            account_name_idx = self._get_field_idx("account_name")

            if client_code_idx != -1 and len(row_data) > client_code_idx:
                client_code = row_data[client_code_idx].strip() if row_data[client_code_idx] else ""
                if client_code:
                    if client_code in client_codes:
                        self.add_error_for_field(row_num, client_code_idx, client_code,
                                      f"Duplicate client code (first occurrence: row {client_codes[client_code]})")
                    else:
                        client_codes[client_code] = row_num

            # Track errors after validation
            errors_after = len([e for e in self.errors if e.severity == Severity.ERROR and e.row_num == row_num])
            warnings_after = len([e for e in self.errors if e.severity == Severity.WARNING and e.row_num == row_num])

            # Record row result
            account_name = row_data[account_name_idx].strip() if account_name_idx != -1 and len(row_data) > account_name_idx else ""
            client_code = row_data[client_code_idx].strip() if client_code_idx != -1 and len(row_data) > client_code_idx else ""
            error_count = errors_after - errors_before
            warning_count = warnings_after - warnings_before

            self.row_results.append(RowValidationResult(
                row_num=row_num,
                account_name=account_name,
                client_code=client_code,
                is_valid=(error_count == 0),
                error_count=error_count,
                warning_count=warning_count
            ))

        return len([e for e in self.errors if e.severity == Severity.ERROR]) == 0, self.errors

    def _read_xlsx(self, file_path: str) -> list[list[str]]:
        """Read data from Excel file, skipping any instructional text above the header"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # Find header row (look for "联交所参与者名称" or "序号" for Shenzhen)
            # Search up to row 30 to handle files with instructions/text above the data
            header_row = None
            for row_idx in range(1, min(30, ws.max_row + 1)):
                for col_idx in range(1, min(10, ws.max_column + 1)):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    # For Shenzhen, first column is 序号, second is 联交所参与者名称
                    # For Shanghai, first column is 联交所参与者名称
                    if cell_val and "联交所参与者名称" in str(cell_val):
                        header_row = row_idx
                        self.header_row = header_row  # Store for reference
                        break
                if header_row:
                    break

            if not header_row:
                self.add_error(0, "File", "file", 0, file_path,
                              "Could not find header row with '联交所参与者名称'. Please ensure the Excel file contains the correct header row.")
                return []

            # Determine number of columns based on exchange type
            # Shenzhen: 38 columns, Shanghai: 42 columns
            max_cols = 39 if self.exchange_type == "SHENZHEN" else 43  # +1 for 1-based Excel indexing

            # Read data rows (skip header row and any empty rows)
            rows = []
            for row_idx in range(header_row + 1, ws.max_row + 1):
                row_data = []
                has_data = False
                for col_idx in range(1, max_cols):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = str(cell.value) if cell.value is not None else ""
                    row_data.append(val)
                    if val and val != "None":
                        has_data = True
                if has_data:
                    rows.append(row_data)

            return rows

        except ImportError:
            self.add_error(0, "File", "file", 0, file_path,
                          "openpyxl library required for Excel files. Install with: pip install openpyxl")
            return []
        except Exception as e:
            self.add_error(0, "File", "file", 0, file_path, f"Error reading Excel file: {str(e)}")
            return []

    def generate_report(self) -> str:
        """Generate a validation report"""
        lines = []
        lines.append("=" * 80)

        # Exchange-specific titles
        if self.exchange_type == "SHENZHEN":
            lines.append("SZSE Programmatic Trading Report Validation Results")
            lines.append("深股通投资者程序化交易信息报告表验证结果")
        else:
            lines.append("SSE Programmatic Trading Report Validation Results")
            lines.append("沪股通投资者程序化交易信息报告表验证结果")

        lines.append("=" * 80)
        lines.append("")

        # Display exchange information
        if self.exchange_type:
            exchange_label = "Shenzhen Stock Exchange (深圳证券交易所)" if self.exchange_type == "SHENZHEN" else "Shanghai Stock Exchange (上海证券交易所)"
            lines.append(f"Exchange: {exchange_label}")
        if self.firm_id:
            lines.append(f"Broker Code (FIRM_ID): {self.firm_id}")
        if self.submission_date:
            lines.append(f"Submission Date: {self.submission_date.strftime('%Y-%m-%d')}")
        if self.exchange_type or self.firm_id or self.submission_date:
            lines.append("")

        errors = [e for e in self.errors if e.severity == Severity.ERROR]
        warnings = [e for e in self.errors if e.severity == Severity.WARNING]

        # Summary
        total_rows = len(self.row_results)
        valid_rows = len([r for r in self.row_results if r.is_valid])
        invalid_rows = total_rows - valid_rows

        lines.append(f"Total Rows Processed: {total_rows}")
        lines.append(f"Valid Rows: {valid_rows}")
        lines.append(f"Invalid Rows: {invalid_rows}")
        lines.append(f"Total Errors: {len(errors)}")
        lines.append(f"Total Warnings: {len(warnings)}")
        lines.append("")

        # Successful validations
        if valid_rows > 0:
            lines.append("-" * 80)
            lines.append("SUCCESSFUL VALIDATIONS:")
            lines.append("-" * 80)
            for result in self.row_results:
                if result.is_valid:
                    context_parts = []
                    if result.account_name:
                        context_parts.append(result.account_name)
                    if result.client_code:
                        context_parts.append(f"BCAN: {result.client_code}")
                    context = " - ".join(context_parts) if context_parts else "Row data"

                    msg = f"✓ Row {result.row_num}: {context}"
                    if result.warning_count > 0:
                        msg += f" (with {result.warning_count} warning(s))"
                    lines.append(msg)
            lines.append("")

        if errors:
            lines.append("-" * 80)
            lines.append("ERRORS:")
            lines.append("-" * 80)
            for e in errors:
                lines.append(str(e))
            lines.append("")

        if warnings:
            lines.append("-" * 80)
            lines.append("WARNINGS:")
            lines.append("-" * 80)
            for w in warnings:
                lines.append(str(w))
            lines.append("")

        if not errors and not warnings:
            lines.append("=" * 80)
            lines.append("✓ ALL VALIDATIONS PASSED!")
            lines.append("=" * 80)

        return "\n".join(lines)


def main():
    """Main entry point"""
    if len(sys.argv) < 2:
        print("Usage: python chinatest.py <excel_file_path>")
        print("       Only Excel files (.xlsx) are supported")
        sys.exit(1)

    file_path = sys.argv[1]
    validator = SSEValidator()
    is_valid, errors = validator.validate_file(file_path)

    print(validator.generate_report())

    sys.exit(0 if is_valid else 1)


if __name__ == "__main__":
    main()

